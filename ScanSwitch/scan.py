import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import os
import nmap
import datetime
import time
start_time = time.time()
wb = openpyxl.load_workbook('Общая таблица по коммутаторам.xlsx')
sheet = wb['Лист1']
sheet.insert_cols(6)
title = sheet['F1']
now =  datetime.datetime.now()
title.value = now.strftime('%d-%m-%Y %H:%M')
title.border = openpyxl.styles.Border(left=Side(border_style='thin', color='FF000000'), 
                                                right=Side(border_style='thin', color='FF000000'), 
                                                top=Side(border_style='thin', color='FF000000'), 
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                )
nm = nmap.PortScanner()
host = ''
x = 2
for row in sheet.iter_rows(min_col=5, max_col=5, min_row=2, max_row=sheet.max_row):
    for cell in row:
        if isinstance(cell.value, str):
            host = cell.value
            nm.scan(hosts=host, arguments='-p T:22 -Pn') 
            state = nm[host]['tcp'][22]['state']
            print('{} : {}'.format(host, state))
            cellF = sheet.cell(x, 6)
            cellF.value = state
            x += 1

# host_list = [(x, nm[x]['tcp'][22]['state']) for x in nm.all_hosts()] 
# for host, state in host_list:
#     stateVal = state
#     hostVal = host
#     # cellD = sheet.cell(x, 6)
#     # cellD.value = stateVal
#     x += 1

for row in sheet.iter_rows(min_col=6, max_col=6, min_row=2, max_row=sheet.max_row):
    for cell in row:
        cell.border = openpyxl.styles.Border(left=Side(border_style='thin', color='FF000000'), 
                                                right=Side(border_style='thin', color='FF000000'), 
                                                top=Side(border_style='thin', color='FF000000'), 
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                )
        if cell.value == 'open':
            cell.fill = openpyxl.styles.PatternFill(start_color='00ff00', fill_type='solid')            
        elif cell.value == 'filtered':
            cell.fill = openpyxl.styles.PatternFill(start_color='ffa500', fill_type='solid')
        else:
            cell.fill = openpyxl.styles.PatternFill(start_color='ff0000', fill_type='solid')
print("--- %s seconds ---" % (time.time() - start_time))
print('Конец')
wb.save('Общая таблица по коммутаторам.xlsx')